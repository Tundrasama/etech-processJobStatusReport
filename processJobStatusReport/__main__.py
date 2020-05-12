import os.path
import re
import sys
from openpyxl import *
from easygui import fileopenbox, msgbox
from datetime import datetime

#100 - TODO: make into separate, new file -- so not important that the file be created
	 # just get the source file, and open a new workbook to write to and save
#TODO: see about uploading file, without API
#TODO: timer to remove complete confirmation message?
#TODO: add time stamp

def main():
	# select dialog
	ret_val = fileopenbox("Please select a file [Make sure the file is closed]", "Select QuickBooks Job Report")
	print(f'Return value is:{ret_val}')
	jobFile=ret_val
	# jobFile = askopenfilename()
	if jobFile == None:
		print ("Must choose a file.")
		sys.exit()
	wb = load_workbook(filename=jobFile)
	destWB = Workbook()
	destPath = os.environ['USERPROFILE'] + '\McGrawAutomation'
	if not os.path.exists(destPath):
		os.makedirs(destPath)
	now=datetime.now().timestamp()
	now=str(now).split('.')[0]
	destFilename=f"{destPath}\JobNamesReformatted-{now}.xlsx"
	destWS = destWB['Sheet']
	ws=wb['Sheet1']
	startRow=2
	maxRow=len(ws['B'])
	records=maxRow-startRow

	for x in range(startRow,maxRow+1):
		customerLine = ws['B'+str(x)].value
		colonCount=len(customerLine)-len(customerLine.replace(':',''))

		if colonCount == 0:
			colonCount=f"{colonCount} - Root - "
		elif colonCount == 1:
			colonCount=f"{colonCount} - Job - "
			customerPattern=r'(.*)\s\((\d{4})\)'
			jobPattern=r'(\d{5})-(.*)'
			for i,parts in enumerate(customerLine.split(':')):
				if i == 0:
					customerInfo=re.search(customerPattern,parts)			
					customer = customerInfo.group(1)
					customerID = customerInfo.group(2)
				elif i == 1:
					jobInfo=re.search(jobPattern,parts)			
					jobID=jobInfo.group(1).strip()
					job=jobInfo.group(2).strip()

			print(f"[{x}] -- [{customer}] -- [{customerID}] -- [{jobID}] -- [{job}] ")
			# ws['G'+str(x)]=f"{jobID}-{job}"
			destWS['A'+str(x-2)]=f"{jobID}- {job}"
		elif colonCount == 2:
			colonCount=f"{colonCount} - Sub-Job - "
			customerPattern=r'(.*)\s\((\d{4})\)'
			jobPattern=r'(\d{5}-\d{3})(.*)'
			for i,parts in enumerate(customerLine.split(':')):
				if i == 0:
					customerInfo=re.search(customerPattern,parts)			
					customer = customerInfo.group(1)
					customerID = customerInfo.group(2)
				elif i == 2:
					jobInfo=re.search(jobPattern,parts)			
					jobID=jobInfo.group(1).strip()
					job=jobInfo.group(2).strip()
			print(f"[{x}] -- [{customer}] -- [{customerID}] -- [{jobID}] -- [{job}] ")
			# ws['G'+str(x)]=f"{jobID}-{job}"
			destWS['A'+str(x-2)]=f"{jobID}- {job}"
		
	try:
		destWB.save(filename=destFilename)
	except:
		msgbox("Please close the file and run again")
		sys.exit()

	wb.close()
	destWB.close()
	msgbox("Process Complete... File located at : " + destFilename)
	os.system(f'start {os.path.realpath(destFilename)}')
	
if __name__ == "__main__":
	main()
